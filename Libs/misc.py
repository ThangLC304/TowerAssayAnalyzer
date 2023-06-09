import tkinter as tk
from tkinter import filedialog
import cv2
from pathlib import Path
import math
import random
import json
import pandas as pd
import os
import openpyxl
import re
from tqdm import tqdm
import numpy as np
import logging
from collections import OrderedDict

logger = logging.getLogger(__name__)

ORDINALS = ['1st', '2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th']
CHARS = [chr(i) for i in range(65, 65+26)]
TESTS_LIST = ['Novel Tank Test', 
              'Shoaling Test', 
              'Mirror Biting Test',
              'Social Interaction Test',
              'Predator Test']


def fill_space_in_name():
    root = tk.Tk()
    root.withdraw()
    file_type = [('Any', '*.*')]
    default_dir = 'Input'
    file_paths = filedialog.askopenfilenames(filetypes=file_type, initialdir=default_dir)
    for file_path in file_paths:
        file_path = Path(file_path)
        file_name = file_path.stem
        file_name = file_name.replace(' ', '_')
        file_name = file_name + file_path.suffix
        file_path.rename(file_path.parent / file_name)
    
    print('Done')


def get_file_path(type, initialdir, mode = 'single'):
    root = tk.Tk()
    root.withdraw()
    file_type = type
    default_dir = initialdir
    if mode == 'single':
        file_path = filedialog.askopenfilename(filetypes=file_type, initialdir=default_dir)
        return file_path
    if mode == 'multiple':
        file_paths = filedialog.askopenfilenames(filetypes=file_type, initialdir=default_dir)
        return file_paths

def load_raw_df(txt_path, sep = "\t"):
    # Read the .txt file into a dataframe
    raw_df = pd.read_csv(txt_path, sep = sep)
    
    # partial_path = Path(*Path(txt_path).parts[-3:]) if len(Path(txt_path).parts) > 3 else Path(txt_path) 
    
    # if len(raw_df) > 0:
    #     print(f'Loaded raw data from ".\{partial_path}"')
    # else:
    #     print(f'No data loaded from ".\{partial_path}"')

    # Remove unnecessary columns
    tanks_list = []
    for col in raw_df.columns:
        if "unnamed" in col.lower() or "prob" in col.lower():
            raw_df.drop(col, axis=1, inplace=True)
        if "x" in col.lower():
            # find the number in the column name
            tank_num = re.findall(r'\d+', col)
            tank_num = int(tank_num[0])
            tanks_list.append(tank_num)
    # print('Tanks found: ', tanks_list)

    return raw_df, tanks_list


def get_image(source = 'video'):

        if source == 'video':
            # Open dialog to choose the .mp4 file
            file_type = [('Video File', '*.mp4')]
            default_dir = 'Input'
            media_path = get_file_path(file_type, default_dir)

            im = get_thumbnail(media_path)
        
        if source == 'image':
            file_type = [('Image File', '*.png')]
            default_dir = 'Input'
            media_path = get_file_path(file_type, default_dir)

            im = cv2.imread(media_path)

        return im, media_path

def get_thumbnail(video_path):
    video_P = Path(video_path)
    cap = cv2.VideoCapture(video_path)
    _, frame = cap.read()
    cap.release()

    # Export frame as a .png file
    image_P = video_P.parent / (video_P.stem + ".png")
    cv2.imwrite(str(image_P), frame)

    im = cv2.imread(str(image_P))
    return im

def get_tanks(im, tank_names = ['Top Left', 'Top Right', 'Bottom Left']):
    example_tanks = {}

    def tank_selector(im, tank_name):
        while True:
            selected_pixel = cv2.selectROI('Selecting tank at position'+str(tank_name), im)
            
            # print out the coordinates of top left and bottom right pixel of the ROI
            print('Tank at position: ', tank_name)
            print('Top left pixel: x = {}, y = {}'.format(selected_pixel[0], selected_pixel[1]))
            print('Bottom right pixel: x = {}, y = {}'.format(selected_pixel[0] + selected_pixel[2], selected_pixel[1] + selected_pixel[3]))
            print('Middle line: y = {}'.format(selected_pixel[1] + selected_pixel[3] / 2))
            tank_d = math.ceil((selected_pixel[2] + selected_pixel[3]) / 2)
            print('Diameter of the tank : ', tank_d)
            print()
            example_tanks[tank_name] = selected_pixel
            cv2.destroyAllWindows()
            break

    for name in tank_names:
        tank_selector(im, name)

    return example_tanks


def color_generator(color_num):

    color_dict = {}
    for i in range(1, color_num + 1):
        color_dict[i] = (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))

    return color_dict

def color_code_loader():
    return {1: [110, 57, 16], 2: [238, 236, 124], 3: [46, 132, 169], 4: [188, 78, 189], 5: [225, 64, 151], 6: [112, 212, 38], 7: [233, 62, 246], 8: [138, 108, 5], 9: [57, 104, 67], 10: [200, 103, 211]}



def draw_points(input_df, image, tank_number, radius = 2, filled_points = [], draw_nan = False, color_rule = 'tank'):
    assert color_rule in ['tank', 'input'], "color_rule must be either 'tank' or 'input'"    
    # color_dict = color_generator(10)
    color_dict = color_code_loader() 

    # color changes based on tank_number
    if color_rule == 'tank':
        color = color_dict[tank_number]
    for i in range(len(input_df)):
        x = input_df['X' + str(tank_number)][i]
        y = input_df['Y' + str(tank_number)][i]
        # change value type to int, if float round to int
        try:
            x = int(x)
        except:
            if draw_nan == False:
                raise ValueError(f"Value x {x} has type {type(x)}")
            else:
                return image
        try:
            y = int(y)
        except:
            if draw_nan == False:
                raise ValueError(f"Value y {y} has type {type(y)}")
            else:
                return image
        if filled_points == []:
            cv2.circle(image, (x, y), radius, color, -1)
        else:
            # if i in filled_points values, change color to black
            if i in filled_points:
                cv2.circle(image, (x, y), radius, (0, 0, 0), -1)
            else:
                cv2.circle(image, (x, y), radius, color, -1)

    return image


def display_coords(input_df, image, window_name, tanks_list, mouse = True, tanks_dict = {}, wait = 0, filled_history = {}, draw_nan = False):

    im = image.copy()

    if filled_history == {}:
        for i in tanks_list:  
            im = draw_points(input_df, im, i, draw_nan = draw_nan)
    else:
        filled_points = {}
        for key in filled_history.keys():
            filled_points[int(key[-1])] = []
        for key, value in filled_history.items():
            if len(value) == 0:
                continue
            else:
                for k, v in value.items():
                    filled_points[int(key[-1])].append(v)

        print(filled_points)
        
        for i in tanks_list:
            im = draw_points(input_df, im, i, filled_points = filled_points[i])

    # display coordinates when mouse hover over the image
    def mouse_callback(event, x, y, flags, param):
        if event == cv2.EVENT_MOUSEMOVE:
            if tanks_dict == {}:
                print('x = {}, y = {}'.format(x, y))
                print()
            else:
                print('Tank {}: x = {}, y = {}'.format(in_which_tank(x, y, tanks_dict), x, y))
                print()
    
    # initial window name
    cv2.namedWindow(window_name)
    # set mouse callback function for window
    cv2.imshow(window_name, im)

    if mouse:
        cv2.setMouseCallback(window_name, mouse_callback, param = tanks_list)

    if wait == 0:
        cv2.waitKey(0)
        cv2.destroyAllWindows()


def in_box(input_x, input_y, box):
    x, y, w, h = box
    return x <= input_x <= x + w and y <= input_y <= y + h
    

def in_which_tank(x, y, tanks_dict):
    for tank_num, tank_box in tanks_dict.items():
        if in_box(x, y, tank_box):
            return tank_num
    return -1


def coord_generator(box, points = 1000):
    x, y, w, h = box
    coords = []
    for i in range(points):
        x_coord = random.randint(x, x+w)
        y_coord = random.randint(y, y+h)
        coords.append((x_coord, y_coord))
    return coords


def mix_up(input_dict, mix_ratio = 0.3):
    # input_dict has 10 keys 
    # keys are 1, 2, 3, 4, 5, 6, 7, 8, 9, 10
    # values are coordinates inside the tanks
    # mix_ratio is the ratio of mixing up the coordinates
    # mix_ratio = 0.3 means 30% of the coordinates will be mixed up

    for i in range(1, len(input_dict)+1):
        if i == 1:
            continue
        else:
            for j in range(1, len(input_dict[i])+1):
                if random.random() < mix_ratio:
                    # swap the coordinates
                    input_dict[i][j-1], input_dict[i-1][j-1] = input_dict[i-1][j-1], input_dict[i][j-1]
    return input_dict


def try_new_name(directory_path, ori_name):
    parent_path = Path(directory_path)
    # json files in parent_path
    json_files = list(parent_path.glob('*.json'))
    sfx = 1
    new_name = ori_name.stem + '_' + str(sfx) + ori_name.suffix
    while True:
        if new_name in json_files:
            sfx += 1
            new_name = ori_name.stem + '_' + str(sfx) + ori_name.suffix
        else:
            break

    return new_name

def load_threshold(threshold_path, *args):
    # load threshold from a json file
    with open(threshold_path, 'r') as f:
        threshold = json.load(f)

    # convert all values to int
    for key, value in threshold.items():
        threshold[key] = int(value)

    return_result = []
    for arg in args:
        if arg not in threshold:
            raise ValueError(f"{arg} is not in the threshold file")
        else:
            return_result.append(threshold[arg])    

    return return_result

def remove_first_row_if_nan(input_df, limitation):
    # if the first row of the dataframe has nan values, remove the entire first row
    while len(input_df) > limitation:
        row_0 = input_df.iloc[0, :]
        if row_0.isnull().values.any():
            input_df = input_df.iloc[1:, :]
            # reset index
            input_df = input_df.reset_index(drop=True)
        else:
            break
    return input_df


def split_array(my_array, sep = 1):
    # get the difference between each number
    diff = np.diff(my_array)
    # get the index of the difference that is larger than sep
    split_index = np.where(diff > sep)[0]
    # split the array into groups
    split_array = np.split(my_array, split_index + 1)
    return split_array



def clean_df(input_df, fill = False, frames = 0, DEBUG = False, remove_nan = True, limitation = 15000): 

    # Remove the initial rows with nan values
    if remove_nan:
        input_df = remove_first_row_if_nan(input_df, limitation)

    # Only take the first frames rows
    if frames == 0:
        pass
    else:
        input_df = input_df.iloc[:frames, :]


    # filled_history = {} # {[filled_value]: [nan_coords]}
    # if fill == True:
    #     # fill nan values with the previous value
    #     for col in input_df.columns:
    #         # initialize the filled_history of each column
    #         filled_history[col] = {}

    #         if DEBUG:
    #             print('In column: ', col, '...')
    #         # find the nan values
    #         nan_coords = np.where(input_df[col].isnull())[0]    # e.g. array([3, 5], dtype=int64)
    #         if DEBUG:
    #             print('Current NaN coordinates: ', nan_coords)
    #         # if there are nan values
    #         if len(nan_coords) == 0:
    #             continue
            
    #         # split nan_coords into groups of continuous numbers (sep = 1)
    #         nan_coords_groups = split_array(nan_coords)

    #         for nan_coords_group in nan_coords_groups:
    #             # find the previous value
    #             prev_value = input_df[col][nan_coords_group[0] - 1]
    #             # # fill the nan values with the previous value
    #             # input_df[col][nan_coords_group] = prev_value
    #             # record the filled values
    #             if prev_value not in filled_history:
    #                 filled_history[col][prev_value] = nan_coords_group
    #             else:
    #                 filled_history[col][prev_value] = np.concatenate((filled_history[col][prev_value], nan_coords_group))
    #             # change the value in filled_history from np.array to list
    #             filled_history[col][prev_value] = filled_history[col][prev_value].tolist()
        
    #     # fill the nan values with the previous value
    #     input_df = input_df.fillna(method='ffill')

    if fill == False:
        return input_df
    
    # Fill the nan values using forward fill and backward fill
    output_df = input_df.fillna(method='ffill')
    output_df = output_df.fillna(method='bfill')

    return output_df


def append_df_to_excel(filename, df, sheet_name='Sheet1', startcol=None, startrow=None, col_sep = 0, row_sep = 0,
                       truncate_sheet=False, DISPLAY = False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        logger.info("Excel file doesn't exist - directly export using df.to_excel")
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startcol=startcol if startcol is not None else 0, 
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name]
        row_0 = ws[1]
        logger.debug(f"Header: {row_0}")

        return
    
    logger.info("Excel file exists - appending using openpyxl")

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    # try to open an existing workbook
    writer.workbook = openpyxl.load_workbook(filename)

    # get the last col in the existing Excel sheet
    # if it was not specified explicitly
    if startcol is None and sheet_name in writer.workbook.sheetnames:
        startcol = writer.workbook[sheet_name].max_column + col_sep

    if startrow is None and sheet_name in writer.workbook.sheetnames:
        startrow = writer.workbook[sheet_name].max_row + row_sep
    
    if startcol is None:
        startcol = 0

    if startrow is None:
        startrow = 0

    logger.debug(f'In file: {os.path.basename(filename)}')
    logger.debug(f'Appended to column: {startcol}, row: {startrow}')

    # row_0 = writer.workbook[sheet_name][1]
    # logger.debug(f"Header: {row_0}")
    
    # remove df headers if they exist
    if startrow != 0:
        # take the first row
        first_row = df.iloc[0].astype(str)
        # check if any cell in the first row contains a letter
        has_letter = first_row.str.contains('[a-zA-Z]').any()
        if has_letter:
            df = df.iloc[1:, :]

    # write the dataframe to the existing sheet
    df.to_excel(writer, sheet_name, startcol=startcol, startrow=startrow, **to_excel_kwargs)

    # close workbook
    writer.close()


def draw_prep(mode = 'multiple', im_path = None, txt_path = None, sep = ","):
    assert mode in ['single', 'multiple'], "mode must be either 'single' or 'multiple'"

    if im_path == None:
        im, _ = get_image(source = 'image')
    else:
        im = cv2.imread(im_path)

    if txt_path == None:
        file_type = [('Txt or csv files', '*.txt;*.csv')]
        default_dir = 'Input'
        txt_paths = get_file_path(file_type, default_dir, mode = mode)
    else:
        txt_paths = txt_path

    final_df = pd.DataFrame()
    tanks_list = []
    # loop through all the txt files
    # read df and merge to final_df
    print("Reading data")
    if mode == 'single':
        final_df, _ = load_raw_df(txt_paths, sep = sep)
        for col in final_df.columns:
            if 'x' in col.lower():
                tank_num = int(col.lower().split('x')[1])
                tanks_list.append(tank_num)
    else:
        for txt_path in tqdm(txt_paths):
            txt_stem = Path(txt_path).stem
            tank_num = int(txt_stem.split('_')[1])
            tanks_list.append(tank_num)
            df, _ = load_raw_df(txt_path)
            columns = [f'X{tank_num}', f'Y{tank_num}']
            df.columns = columns
            final_df = pd.concat([final_df, df], axis = 1)


        print("Cleaning data")
        final_df, filled_history = clean_df(final_df, fill = True, frames = 3000)

        print('Exported filled_history to .json file')
        # indent = 4 for pretty print
        json_path = r'D:\Code\TowerAssayAnalyzer\History\manual_filled_history.json'
        with open(json_path, 'w') as f:
            json.dump(filled_history, f, indent = 4)

    print('Tanks list', tanks_list)
    print('Final DF')
    print(final_df)

    return final_df, im, tanks_list

def draw_trajectories(final_df, im, tanks_list, until = 3000, mouse = False, wait = 0, filled_history = {}, draw_nan = False):
    final_df = final_df.iloc[:until, :]
    display_coords(final_df, im, window_name = "Display trajectories", mouse = mouse, tanks_list = tanks_list, wait = wait, filled_history = filled_history, draw_nan = draw_nan)


def open_dir(output_dir):
    os.startfile(output_dir)

def create_messagebox(root, title, output_dir, button=True):

    message = f'The summary stats are exported to {output_dir}'
    # pop-up a message to notify that the file has been exported
    messagebox = tk.Toplevel(root)
    messagebox.title(title)
    # flexibly set the size of messagebox
    messagebox.geometry('+{}+{}'.format(root.winfo_x() + 500, root.winfo_y() + 250))
    messagebox.configure(bg="white")
    label = tk.Label(messagebox, text=message, font=('Arial', 16, 'bold'))
    label.grid(row=0, column=0, padx=10, pady=20)
    if button:
        go_button = tk.Button(messagebox, text='Go to Output directory', font=('Arial', 16, 'bold'), command=lambda: open_dir(output_dir))
        go_button.grid(row=1, column=0, padx=10, pady=20)
        go_button.configure(bg = 'dark green', fg = 'white')
    # change the messagebox size to fit the label
    messagebox.update()
    messagebox.after(3000, messagebox.destroy)


def get_sheet_names(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    return wb.sheetnames

def find_existed_batches(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    
    max_row_dict = {}
    existed_batches = {}

    # get the first sheet
    for worksheet in workbook.worksheets:
        max_row = worksheet.max_row
        max_row_dict[worksheet.title] = max_row

        # find unique values of first column
        first_col = [worksheet.cell(row=i, column=1).value for i in range(2, max_row+1)]
        first_col = [i for i in first_col if i != None]
        first_col = list(set(first_col))
        for batch in first_col:
            if batch == None:
                continue
            elif batch not in existed_batches:
                existed_batches[batch] = 1
            else:
                existed_batches[batch] += 1

    return existed_batches


def merge_cells(file_path, col_name = 'Shoaling Area', cell_step=3, inplace = True):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    for worksheet in workbook.worksheets:
        # Find the column index for the "Shoaling Area" header
        shoaling_area_col = None
        for col_idx in range(1, worksheet.max_column+1):
            header = worksheet.cell(row=1, column=col_idx).value
            if header and col_name in header:
                shoaling_area_col = col_idx
                break

        if shoaling_area_col is None:
            print("Column not found.")
        else:
            # Merge every next 3 rows of the Shoaling Area column
            for row_idx in range(2, worksheet.max_row+1, cell_step):
                value = worksheet.cell(row=row_idx, column=shoaling_area_col).value
                # print(value)
                if value is not None:
                    # Merge the current row with the next 2 rows
                    worksheet.merge_cells(start_row=row_idx, start_column=shoaling_area_col, end_row=row_idx+2, end_column=shoaling_area_col)
                
                # align the merged cell, horizontal and vertical center
                worksheet.cell(row=row_idx, column=shoaling_area_col).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
    # define output_path
    if inplace == False:
        output_path = file_path[:-5] + '_merged.xlsx'        
    else:
        output_path = file_path    

    # Save the modified workbook
    workbook.save(filename=output_path)

def excel_polish(file_path, batch_num, cell_step=10, treatment = None, inplace=True):

    print("Polishing excel file...")

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)


    # Adjust the column widths
    # Loop through each sheet in the workbook
    for sheet_name in workbook.sheetnames:

        if "analysis" in sheet_name.lower():
            continue
        # Select the sheet
        sheet = workbook[sheet_name]
        
        # Loop through each column in the sheet
        for col in sheet.columns:
            # Set the width of the column to 17.00 (160 pixels)
            sheet.column_dimensions[col[0].column_letter].width = 17.00
        
        # Enable text wrapping for the header row
        for cell in sheet[1]:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')


    # find column index with header = "Fish ID"
    fish_id_col = None
    for worksheet in workbook.worksheets:
        for col_idx in range(1, worksheet.max_column+1):
            header = worksheet.cell(row=1, column=col_idx).value
            if header and "Fish ID" in header:
                fish_id_col = col_idx
                logger.debug("Found fish_id_col at column %d", fish_id_col)
                break

    # find row index of cell with "Fish 1" in fish_id_col
    fish1_rows = {}
    for worksheet in workbook.worksheets:
        fish1_rows[worksheet.title] = []

        for row_idx in range(1, worksheet.max_row+1):
            fish_id = worksheet.cell(row=row_idx, column=fish_id_col).value
            if fish_id and fish_id == "Fish 1":
                fish1_rows[worksheet.title].append(row_idx)

    def merge_cells(ws, merge_column, merge_value, start_row, end_row):
        # merge the next {cell_step} rows
        ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=end_row, end_column=merge_column)
        # put merge_value in the merged cell
        ws.cell(row=start_row, column=merge_column).value = merge_value
        # font 12, Calibri, Bold
        ws.cell(row=start_row, column=merge_column).font = openpyxl.styles.Font(name='Calibri', size=12, bold=True)
        ws.cell(row=start_row, column=merge_column).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    BATCH_NUM_COL = 1
    TREATMENT_COL = 2
    BATCH_TEXT = f"Batch {batch_num}"

    # Add a separator line between each update
    # Iterate in fish_id_col, from fish1_rows[worksheet.title][-1] to the end of the column
    for worksheet in workbook.worksheets:
        logger.debug("In worksheet %s", worksheet.title)
        start_row = fish1_rows[worksheet.title][-1] # Because we modify only the last update
        last_row = start_row # In case there's only 1 fish
        logger.debug(f"Total row of the worksheet: {worksheet.max_row}")
        logger.debug(f"Content of the last row: {worksheet.cell(row=worksheet.max_row, column=fish_id_col).value}")
        logger.debug(f"Start row: {start_row}, last row: {last_row}")
        for row_idx in range(fish1_rows[worksheet.title][-1], worksheet.max_row+1):
            cell_value = worksheet.cell(row=row_idx, column=fish_id_col).value
            logger.debug(f"Cell value: {cell_value}")
            fish_id = int(cell_value.split()[1])
            logger.debug(f"Fish id: {fish_id}")
            next_cell_value = worksheet.cell(row=row_idx+1, column=fish_id_col).value
            logger.debug(f"Next cell value: {next_cell_value}")
            try:
                next_fish_id = int(next_cell_value.split()[1])
            except:
                next_fish_id = -1
            logger.debug("Next fish id: %d", next_fish_id)
            if next_fish_id - fish_id != 1:
                logger.debug("Found a separator at row %d", row_idx)
                last_row = row_idx
                try:
                    merge_cells(ws = worksheet,
                            merge_column = BATCH_NUM_COL,
                            merge_value = BATCH_TEXT,
                            start_row = start_row,
                            end_row = last_row)
                    logger.debug("Merged cells in worksheet %s", worksheet.title)
                except:
                    logger.warning("Cannot merge cells in worksheet %s", worksheet.title)
                
                if treatment is not None:
                    logger.debug("Also merge treatment column")
                    merge_cells(ws = worksheet,
                                merge_column = TREATMENT_COL,
                                merge_value = treatment,
                                start_row = start_row,
                                end_row = last_row)

                # Check if the cell in last_row+1 is "Separator"?
                if worksheet.cell(row=last_row+1, column=fish_id_col).value == "Separator":
                    logger.debug("Already had a separator at row %d", last_row+1)
                    break
                # insert a blank row with fish_id = "Separator" at the end of the last update
                try:
                    worksheet.insert_rows(last_row+1)
                    worksheet.cell(row=last_row+1, column=fish_id_col).value = "Separator"
                    logger.debug("Inserted a separator row at row %d", last_row+1)
                except:
                    logger.warning("Cannot insert a separator row at row %d", last_row+1)
                break

        print(f"In worksheet {worksheet.title}, last update has {last_row-start_row+1} fish, start from row {start_row} to row {last_row}")

    # define output_path
    if inplace == False:
        output_path = file_path.replace(str(Path(file_path).suffix),'_merged.xlsx')      
    else:
        output_path = file_path    

    # Save the modified workbook
    workbook.save(filename=output_path)


def sort_paths_by_parent(paths):

    def sort_key(path):
        name_parts = path.parent.name.split('-')
        if len(name_parts) >= 2:
            try:
                primary_key = int(name_parts[0].strip())
            except ValueError:
                primary_key = int(name_parts[0].strip())
            secondary_key = int(name_parts[1].strip())
        else:
            primary_key = int(name_parts[0].strip())
            secondary_key = 0
        return (primary_key, secondary_key)
    
    return sorted(paths, key=sort_key)


def hyploader(hyp_path):

    with open(hyp_path, 'r') as file:
        data = json.load(file)
        
    # convert values to int or float
    for key, value in data.items():
        if key == "CONVERSION RATE":
            data[key] = float(value)
        elif key in ["FRAME RATE", "DURATION", "SEGMENT DURATION", "ZONE WIDTH"]:
            data[key] = int(float(value))
        else:
            for fish_num, fish_data in value.items():
                if isinstance(fish_data, list):
                    for i, item in enumerate(fish_data):
                        fish_data[i] = int(float(item))
                else:
                    data[key][fish_num] = int(float(fish_data))

    def zone_calculator(target_name, target_data, reverse=False):
        zone_name = target_name + " ZONE"
        target_data[zone_name] = {}
        n = 1 if reverse == False else -1
        for fish_num, fish_data in target_data[target_name].items():
            m = -1 if fish_data[1] == 0 else 1
            target_data[zone_name][fish_num] = [fish_data[0] + m * n * target_data["ZONE WIDTH"] *  target_data["CONVERSION RATE"], fish_data[1]]

        return target_data

    if "MIRROR" in data.keys():
        data = zone_calculator("MIRROR", data, reverse=True)
    elif "SEPARATOR" in data.keys():
        data = zone_calculator("SEPARATOR", data)

    return data

def get_static_dir(project_dir, batch_num, treatment_count):

    if not os.path.exists(project_dir):
        return
    
    p_dir = Path(project_dir)

    temp_paths = []

    for i in range(treatment_count):
        treatment_char = chr(ord('A') + i)
        temp_paths.append(p_dir / "static" / f"Batch {batch_num}" / treatment_char)

    return temp_paths

def natural_sort_keys(path):
    dir_name, file_name = os.path.split(path)
    return [int(text) if text.isdigit() else text.lower() for text in re.split('(\d+)', file_name)]


def check_trajectories_dir(project_dir, current_test, current_treatment, current_batch):

    logger.info(f"Checking trajectories dir for Project:{project_dir}, {current_test}, {current_treatment}, {current_batch}")

    # current_test = Novel Tank Test
    # current_batch = Batch 1
    # current_treatment = Treatment A

    batch_num = current_batch.split()[1]
    # change batch_num to ordinal
    batch_num = int(batch_num)
    batch_num = ORDINALS[batch_num-1]

    treatment_char = current_treatment

    project_dir_path = Path(project_dir)

    # use glob to find directory with name include current_test

    test_dirs = project_dir_path.glob(f"*{current_test}*")
    test_dir = list(test_dirs)[0]
    logger.debug(f"Test dir pattern: *{current_test}* => test_dir found: {test_dir}")


    pattern = f"*{treatment_char} -*({batch_num} Batch)"
    treatment_dirs = test_dir.glob(pattern)
    treatment_dir = list(treatment_dirs)[0]
    logger.debug(f"Treatment dir pattern: {pattern} => treatment_dir found: {treatment_dir}")


    # find all directories inside treatment_dir, no pattern, just directories

    trajectories_dirs = [x for x in treatment_dir.iterdir() if x.is_dir()]
    logger.debug(f"Total {len(trajectories_dirs)} trajectories_dirs found in {treatment_dir}")

    checker_dict = {}
    # go into each trajectories_dir, find .txt file, if existed, set checker_dict[trajectories_dir] = True
    for trajectories_dir in trajectories_dirs:
        txt_files = trajectories_dir.glob("*.txt")

        # take out the project_dir from trajectories_dir

        checker_key = str(trajectories_dir.relative_to(project_dir_path))

        if len(list(txt_files)) > 0:
            checker_dict[checker_key] = True
        else:
            checker_dict[checker_key] = False

    if len(checker_dict):
        logger.info("Checked successfully!")
    else:
        logger.info("No trajectories folder found!")

    # Create a new dictionary with keys sorted in natural order
    checker_dict = OrderedDict(sorted(checker_dict.items(), key=lambda item: natural_sort_keys(item[0])))
    logger.debug("Rearranged checker_dict to natural order")

    return checker_dict

    
def nanlize(input_dict, test_index):
    if test_index == 0:
        for time_period, end_points in input_dict.items():
            for end_point in end_points.keys():
                input_dict[time_period][end_point][1] = np.nan 
        
    else:
        for end_point in input_dict.keys():
            input_dict[end_point][1] = np.nan

    return input_dict

def get_treatment_name_from_index(treatment_index, treatment_name_list):
    # turn treatment_index = A to number = 0
    treatment_index = ord(treatment_index) - ord('A')
    return treatment_name_list[treatment_index]

def get_treatment_index_from_name(treatment_name, treatment_name_list):
    # get number from treatment_name_list
    treatment_name_list.index(treatment_name)
    # turn it to char
    return chr(ord('A') + treatment_name_list.index(treatment_name))


    



